import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import collections
import pandas as pd
import numpy as np
import sqlite3
import uuid

EXCEL_FILE = "depo.xlsx"
DB_FILE = "depo.db"
SHEETS = {
    "urunler": "Məhsullar",
    "hareketler": "Hareketler",
    "kategoriler": "Kategoriler"
}

DARK_COLORS = {
    "bg": "#0f172a",        # Modern slate dark
    "sidebar": "#1e293b",   # Slightly lighter slate
    "card": "#1e293b",
    "accent": "#38bdf8",    # Sky blue
    "accent2": "#10b981",   # Emerald green
    "danger": "#ef4444",    # Red
    "warning": "#f59e0b",   # Amber
    "text": "#f8fafc",      # Slate 50
    "text_dim": "#94a3b8",  # Slate 400
    "border": "#334155",    # Slate 700
    "header": "#1e293b",
    "row_alt": "#1e293b",
    "green_bg": "#064e3b",
    "red_bg": "#7f1d1d",
}

LIGHT_COLORS = {
    "bg": "#f1f5f9",
    "sidebar": "#ffffff",
    "card": "#ffffff",
    "accent": "#0284c7",
    "accent2": "#059669",
    "danger": "#dc2626",
    "warning": "#d97706",
    "text": "#0f172a",
    "text_dim": "#64748b",
    "border": "#e2e8f0",
    "header": "#ffffff",
    "row_alt": "#f8fafc",
    "green_bg": "#dcfce7",
    "red_bg": "#fee2e2",
}

COLORS = LIGHT_COLORS  # Default to light theme



def init_excel():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A252F")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style='thin', color='2C3E50'),
        right=Side(style='thin', color='2C3E50'),
        top=Side(style='thin', color='2C3E50'),
        bottom=Side(style='thin', color='2C3E50')
    )

    # Kategoriler
    if SHEETS["kategoriler"] not in wb.sheetnames:
        ws = wb.create_sheet(SHEETS["kategoriler"])
        headers = ["Kategori ID", "Kategori Adı"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
        for row in [["K001", "Elektronik"], ["K002", "Gıda"], ["K003", "Kıyafet"], ["K004", "Mobilya"], ["K005", "Diğer"]]:
            ws.append(row)
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25

    # Məhsullar — yeni sütunlar
    if SHEETS["urunler"] not in wb.sheetnames:
        ws = wb.create_sheet(SHEETS["urunler"])
        headers = [
            "Məhsul",
            "Gəliş sayı",
            "Satılan",
            "Əldə olan",
            "Alış qiyməti",
            "Satış qiyməti",
            "Gəlir",
            "Maya dərəri",
            "Son Yeniləmə"
        ]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
        widths = [25, 14, 12, 14, 18, 18, 16, 16, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Hareketler
    if SHEETS["hareketler"] not in wb.sheetnames:
        ws = wb.create_sheet(SHEETS["hareketler"])
        headers = ["Hareket ID", "Tarih", "Məhsul", "Hareket Türü", "Miktar", "Birim Qiymət", "Açıklama", "Kayıt Eden"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
        widths = [14, 20, 25, 15, 10, 15, 30, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_FILE)


def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            name TEXT UNIQUE,
            gelis_sayi INTEGER,
            satilan INTEGER,
            elde_olan INTEGER,
            alis_qiymeti REAL,
            satis_qiymeti REAL,
            gelir REAL,
            maya REAL,
            son_yenileme TEXT,
            id TEXT PRIMARY KEY,
            is_active INTEGER DEFAULT 1
        )
    ''')
    # Ensure new columns exist for older versions
    try:
        cursor.execute('ALTER TABLE products ADD COLUMN id TEXT')
    except Exception:
        pass
    try:
        cursor.execute('ALTER TABLE products ADD COLUMN is_active INTEGER DEFAULT 1')
    except Exception:
        pass
    
    # Movements Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS movements (
            id TEXT PRIMARY KEY,
            tarih TEXT,
            product_name TEXT,
            type TEXT,
            miktar INTEGER,
            unit_price REAL,
            description TEXT,
            user TEXT
        )
    ''')
    
    # Categories Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id TEXT PRIMARY KEY,
            name TEXT
        )
    ''')
    
    conn.commit()
    conn.close()


def sync_excel_to_db():
    """Excel-dəki məlumatları SQLite-a köçürür."""
    init_db()
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # Məhsulları köçür
        products = []
        if SHEETS["urunler"] in wb.sheetnames:
            ws_u = wb[SHEETS["urunler"]]
            for row in ws_u.iter_rows(min_row=2, values_only=True):
                if row[0]: products.append(row[:9])
            
        # Hərəkətləri köçür
        movements = []
        if SHEETS["hareketler"] in wb.sheetnames:
            ws_h = wb[SHEETS["hareketler"]]
            for row in ws_h.iter_rows(min_row=2, values_only=True):
                if row[0]: movements.append(row[:8])

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM products")
        # After sync, mark all as active
        product_rows = []
        for row in products:
            product_rows.append((*row, str(uuid.uuid4()), 1))
        cursor.executemany("INSERT OR REPLACE INTO products (name, gelis_sayi, satilan, elde_olan, alis_qiymeti, satis_qiymeti, gelir, maya, son_yenileme, id, is_active) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                       product_rows)
        # Ensure is_active flag for all
        cursor.execute("UPDATE products SET is_active=1")
            
        cursor.execute("DELETE FROM movements")
        cursor.executemany("INSERT OR REPLACE INTO movements VALUES (?,?,?,?,?,?,?,?)", movements)
            
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Sinxronizasiya xətası: {e}")


def load_products():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products WHERE is_active=1")
    products = [list(row) for row in cursor.fetchall()]
    conn.close()
    return products


def load_movements():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM movements")
    movements = [list(row) for row in cursor.fetchall()]
    conn.close()
    return movements


def load_categories():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM categories")
    cats = [row[0] for row in cursor.fetchall()]
    conn.close()
    if not cats: # Fallback to Excel if DB is empty
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEETS["kategoriler"]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]: cats.append(row[1])
    return cats


def get_analytics_data():
    """
    SQLite kullanarak ətraflı analiz məlumatlarını qaytarır.
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        
        # 1. Ümumi Stok Dəyəri
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(elde_olan * alis_qiymeti) FROM products")
        total_stock_value = cursor.fetchone()[0] or 0
        
        # 2. Potensial Qazanc
        cursor.execute("SELECT SUM(elde_olan * (satis_qiymeti - alis_qiymeti)) FROM products")
        potential_profit = cursor.fetchone()[0] or 0
        
        # 3. Ən çox satılan 5 məhsul
        cursor.execute("SELECT name, satilan FROM products ORDER BY satilan DESC LIMIT 5")
        top_selling = cursor.fetchall()
        
        # 4. Aylıq Analiz (Satış və Xərc)
        # SQLite-da Pandas-a ehtiyac olmadan birbaşa SQL ilə də hesablaya bilərik amma Pandas rahatdır
        df_movements = pd.read_sql_query("SELECT * FROM movements", conn)
        df_movements["tarih"] = pd.to_datetime(df_movements["tarih"], format="%d.%m.%Y %H:%M", errors='coerce')
        df_movements["Total"] = df_movements["miktar"] * df_movements["unit_price"]
        
        monthly_report = {}
        if not df_movements.empty:
            monthly_stats = df_movements.groupby([df_movements["tarih"].dt.to_period("M"), "type"])["Total"].sum().unstack().fillna(0)
            for period in monthly_stats.index:
                p_str = str(period)
                monthly_report[p_str] = {
                    "income": monthly_stats.loc[period].get("Satış", 0),
                    "expense": monthly_stats.loc[period].get("Giriş", 0)
                }
        
        cursor.execute("SELECT COUNT(*) FROM products")
        product_count = cursor.fetchone()[0]
        
        conn.close()
        return {
            "total_stock_value": total_stock_value,
            "potential_profit": potential_profit,
            "top_selling": top_selling,
            "monthly_report": monthly_report,
            "product_count": product_count
        }
    except Exception as e:
        print(f"Analiz xətası: {e}")
        return None


def get_filtered_analysis_data(months=1):
    """
    Müəyyən ay sayına görə filtrlənmiş analiz məlumatlarını qaytarır.
    months: 1, 3, 6 və ya 12 (illik)
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        df = pd.read_sql_query("SELECT * FROM movements", conn)
        conn.close()

        if df.empty:
            return None

        df["tarih"] = pd.to_datetime(df["tarih"], format="%d.%m.%Y %H:%M", errors='coerce')
        df["Total"] = pd.to_numeric(df["miktar"], errors='coerce') * pd.to_numeric(df["unit_price"], errors='coerce')

        # Ensure numeric Total column
        df["Total"] = pd.to_numeric(df["Total"], errors='coerce')

        # Tarix filteri
        cutoff = datetime.now() - pd.DateOffset(months=months)
        df_filtered = df[df["tarih"] >= cutoff].copy()

        if df_filtered.empty:
            return {"monthly": {}, "product_profit": [], "total_income": 0, "total_expense": 0}

        # Aylıq gəlir/xərc
        monthly_stats = df_filtered.groupby(
            [df_filtered["tarih"].dt.to_period("M"), "type"]
        )["Total"].sum().unstack(fill_value=0)

        monthly = {}
        for period in monthly_stats.index:
            p_str = str(period)
            monthly[p_str] = {
                "income": float(monthly_stats.loc[period].get("Satış", 0)),
                "expense": float(monthly_stats.loc[period].get("Giriş", 0))
            }

        # Məhsul üzrə mənfəət (Satış gəliri - Giriş xərci)
        df_sales = df_filtered[df_filtered["type"] == "Satış"].groupby("product_name")["Total"].sum()
        df_cost  = df_filtered[df_filtered["type"] == "Giriş"].groupby("product_name")["Total"].sum()
        profit_df = df_sales.subtract(df_cost, fill_value=0).reset_index()
        profit_df.columns = ["product", "profit"]
        profit_df = profit_df.sort_values("profit", ascending=False).head(10)
        product_profit = list(zip(profit_df["product"], profit_df["profit"]))

        total_income  = float(df_filtered[df_filtered["type"] == "Satış"]["Total"].sum())
        total_expense = float(df_filtered[df_filtered["type"] == "Giriş"]["Total"].sum())

        return {
            "monthly": monthly,
            "product_profit": product_profit,
            "total_income": total_income,
            "total_expense": total_expense
        }
    except Exception as e:
        print(f"Filtrli analiz xətası: {e}")
        return None


def save_product(data):
    """
    data = [name, gelis_sayi, satilan, elde_olan, alis, satis, gelir, maya, date]
    """
    # 1. SQLite-a yaz
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    # Gelir və Maya formulalarını rəqəmə çevirək (Excel-dən fərqli olaraq SQLite formul saxlamır)
    data_list = list(data)
    data_list[6] = float(data_list[2] or 0) * float(data_list[5] or 0) # Gelir
    data_list[7] = float(data_list[1] or 0) * float(data_list[4] or 0) # Maya
    cursor.execute("INSERT INTO products (name, gelis_sayi, satilan, elde_olan, alis_qiymeti, satis_qiymeti, gelir, maya, son_yenileme, is_active) VALUES (?,?,?,?,?,?,?,?,?,1)", data_list)
    conn.commit()
    conn.close()

    # 2. Excel-ə yaz
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEETS["urunler"]]
    next_row = ws.max_row + 1
    ws.append(data)
    ws.cell(row=next_row, column=7).value = f"=C{next_row}*F{next_row}"
    ws.cell(row=next_row, column=8).value = f"=B{next_row}*E{next_row}"
    wb.save(EXCEL_FILE)


def update_product_stock(product_name, delta, hareket_turu, aciklama, kayit_eden):
    # 1. SQLite-da məlumatları yoxla və yenilə
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products WHERE name=?", (product_name,))
    p = cursor.fetchone()
    
    if not p:
        conn.close()
        return False, "Məhsul tapılmadı!"

    # p: 0=name, 1=gelis, 2=satilan, 3=elde, 4=alis, 5=satis, 6=gelir, 7=maya, 8=date
    gelis = p[1] or 0
    satilan = p[2] or 0
    elde = p[3] or 0
    alis = p[4] or 0
    satis = p[5] or 0

    if hareket_turu == "Giriş":
        new_gelis = gelis + abs(delta)
        new_elde = elde + abs(delta)
        new_satilan = satilan
    elif hareket_turu == "Satış":
        new_satilan = satilan + abs(delta)
        new_elde = elde - abs(delta)
        new_gelis = gelis
        if new_elde < 0:
            conn.close()
            return False, "Kifayət qədər stok yoxdur!"
    
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    new_gelir = new_satilan * satis
    new_maya = new_gelis * alis
    
    cursor.execute("""
        UPDATE products 
        SET gelis_sayi=?, satilan=?, elde_olan=?, gelir=?, maya=?, son_yenileme=?
        WHERE name=?
    """, (new_gelis, new_satilan, new_elde, new_gelir, new_maya, now, product_name))
    
    # Hərəkəti qeyd et
    cursor.execute("SELECT COUNT(*) FROM movements")
    m_count = cursor.fetchone()[0]
    hareket_id = f"H{m_count+1:04d}"
    unit_price = satis if hareket_turu == "Satış" else alis
    cursor.execute("INSERT INTO movements VALUES (?,?,?,?,?,?,?,?)",
                   (hareket_id, now, product_name, hareket_turu, abs(delta), unit_price, aciklama, kayit_eden))
    
    conn.commit()
    conn.close()

    # 2. Excel-i yenilə (Backup)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_u = wb[SHEETS["urunler"]]
    for row in ws_u.iter_rows(min_row=2):
        if row[0].value == product_name:
            r = row[0].row
            ws_u.cell(row=r, column=2).value = new_gelis
            ws_u.cell(row=r, column=3).value = new_satilan
            ws_u.cell(row=r, column=4).value = new_elde
            ws_u.cell(row=r, column=7).value = f"=C{r}*F{r}"
            ws_u.cell(row=r, column=8).value = f"=B{r}*E{r}"
            ws_u.cell(row=r, column=9).value = now
            break
    
    ws_h = wb[SHEETS["hareketler"]]
    ws_h.append([hareket_id, now, product_name, hareket_turu, abs(delta), unit_price, aciklama, kayit_eden])
    wb.save(EXCEL_FILE)
    
    return True, "Əməliyyat uğurlu oldu!"


def delete_product(product_name):
    # 1. SQLite-dan silmek yerine arşivle (is_active=0)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET is_active=0 WHERE name=?", (product_name,))
    conn.commit()
    conn.close()

    # 2. Excel-dən sil
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEETS["urunler"]]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_name:
            ws.delete_rows(row[0].row)
            wb.save(EXCEL_FILE)
            return True
    return False


# ─────────────────────────── UI ───────────────────────────

class DepoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📦 Anbar PRO - İdarəetmə Sistemi")
        self.root.geometry("1200x800")
        self.root.configure(bg=COLORS["bg"])
        init_excel()
        sync_excel_to_db() # SQLite-ı Excel ilə sinxron et
        self.current_user = "Admin"
        self.is_dark = False
        self.build_ui()
        self.navigate("dashboard")

    def navigate(self, key):
        self.current_page = key
        self.set_active_nav(key)
        
        pages = {
            "dashboard": self.show_dashboard,
            "urunler": self.show_products,
            "satis": lambda: self.show_movement("Satış"),
            "hareketler": self.show_movements,
            "maliye": self.show_financials,
            "kritik": self.show_critical,
            "analiz": self.show_analysis,
        }
        pages[key]()

    def toggle_theme(self):
        global COLORS
        self.is_dark = not self.is_dark
        COLORS = DARK_COLORS if self.is_dark else LIGHT_COLORS
        self.root.configure(bg=COLORS["bg"])
        self.sidebar.configure(bg=COLORS["sidebar"])
        self.main.configure(bg=COLORS["bg"])
        self.content.configure(bg=COLORS["bg"])
        
        for widget in self.root.winfo_children():
            widget.destroy()
        self.build_ui()
        self.navigate(self.current_page)

    def build_ui(self):
        # Sidebar
        self.sidebar = tk.Frame(self.root, bg=COLORS["sidebar"], width=240)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # Logo / Brand
        logo_frame = tk.Frame(self.sidebar, bg=COLORS["sidebar"], height=80)
        logo_frame.pack(fill="x", pady=10)
        tk.Label(logo_frame, text="📦", bg=COLORS["sidebar"], fg=COLORS["accent"],
                 font=("Segoe UI", 24)).pack()
        tk.Label(logo_frame, text="ANBAR PRO", bg=COLORS["sidebar"],
                 fg=COLORS["text"], font=("Segoe UI", 12, "bold")).pack()

        tk.Frame(self.sidebar, bg=COLORS["border"], height=1).pack(fill="x", padx=20, pady=10)

        self.nav_buttons = {}
        nav_items = [
            ("🏠  Dashboard", "dashboard"),
            ("📋  Məhsul Siyahısı", "urunler"),
            ("📊  Hərəkətlər", "hareketler"),
            ("💰  Maliyyə", "maliye"),
            ("⚠️  Kritik Stok", "kritik"),
            ("📈  Analiz", "analiz"),
        ]
        for label, key in nav_items:
            btn = tk.Button(self.sidebar, text=f"  {label}", bg=COLORS["sidebar"],
                            fg=COLORS["text"], font=("Segoe UI", 10),
                            bd=0, padx=20, pady=12, anchor="w",
                            cursor="hand2", activebackground=COLORS["accent"],
                            activeforeground="white",
                            command=lambda k=key: self.navigate(k))
            btn.pack(fill="x", padx=10, pady=2)
            self.nav_buttons[key] = btn

        user_frame = tk.Frame(self.sidebar, bg=COLORS["sidebar"])
        user_frame.pack(side="bottom", fill="x", padx=10, pady=20)
        tk.Frame(user_frame, bg=COLORS["border"], height=1).pack(fill="x", pady=(0, 10))
        tk.Label(user_frame, text=f"👤 {self.current_user}", bg=COLORS["sidebar"],
                 fg=COLORS["text_dim"], font=("Segoe UI", 9)).pack()

        self.main = tk.Frame(self.root, bg=COLORS["bg"])
        self.main.pack(side="left", fill="both", expand=True)

        topbar = tk.Frame(self.main, bg=COLORS["header"], height=55)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        self.page_title = tk.Label(topbar, text="Dashboard", bg=COLORS["header"],
                                    fg=COLORS["text"], font=("Segoe UI", 15, "bold"))
        self.page_title.pack(side="left", padx=20, pady=12)

        self.theme_btn = tk.Button(topbar, text="🌓 Tema Dəyiş", bg=COLORS["card"],
                                   fg=COLORS["text"], font=("Segoe UI", 9), bd=0,
                                   padx=12, pady=4, cursor="hand2",
                                   command=self.toggle_theme)
        self.theme_btn.pack(side="right", padx=5, pady=12)

        self.refresh_btn = tk.Button(topbar, text="🔄 Yenilə", bg=COLORS["accent"],
                                      fg="white", font=("Segoe UI", 9), bd=0,
                                      padx=12, pady=4, cursor="hand2",
                                      command=self.refresh_current)
        self.refresh_btn.pack(side="right", padx=15, pady=12)

        self.content = tk.Frame(self.main, bg=COLORS["bg"])
        self.content.pack(fill="both", expand=True, padx=15, pady=10)

        self.current_page = "dashboard"

    def clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def set_active_nav(self, key):
        for k, btn in self.nav_buttons.items():
            btn.config(bg=COLORS["sidebar"])
        if key in self.nav_buttons:
            self.nav_buttons[key].config(bg=COLORS["accent"])

    def refresh_current(self):
        self.navigate(self.current_page)

    def make_card(self, parent, title, value, color, icon):
        card = tk.Frame(parent, bg=COLORS["card"], bd=0)
        # Gradient effect or border
        top_border = tk.Frame(card, bg=color, height=3)
        top_border.pack(side="top", fill="x")
        
        inner = tk.Frame(card, bg=COLORS["card"], padx=20, pady=20)
        inner.pack(fill="both", expand=True)
        
        header = tk.Frame(inner, bg=COLORS["card"])
        header.pack(fill="x")
        
        tk.Label(header, text=icon, bg=COLORS["card"], fg=color,
                 font=("Segoe UI", 16)).pack(side="left")
        tk.Label(header, text=title, bg=COLORS["card"],
                 fg=COLORS["text_dim"], font=("Segoe UI", 10)).pack(side="left", padx=10)
                 
        tk.Label(inner, text=str(value), bg=COLORS["card"],
                 fg=COLORS["text"], font=("Segoe UI", 20, "bold")).pack(anchor="w", pady=(10, 0))
        return card

    def make_table(self, parent, columns, rows, col_widths=None):
        frame = tk.Frame(parent, bg=COLORS["bg"])
        frame.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                         background=COLORS["card"],
                         foreground=COLORS["text"],
                         fieldbackground=COLORS["card"],
                         rowheight=30,
                         font=("Segoe UI", 9))
        style.configure("Custom.Treeview.Heading",
                         background=COLORS["header"],
                         foreground=COLORS["text"],
                         font=("Segoe UI", 9, "bold"),
                         relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "white")])

        tree = ttk.Treeview(frame, columns=columns, show="headings",
                             yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                             style="Custom.Treeview")
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        tree.pack(fill="both", expand=True)

        for i, col in enumerate(columns):
            w = col_widths[i] if col_widths else 120
            tree.heading(col, text=col)
            tree.column(col, width=w, minwidth=60, anchor="center")

        for i, row in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            tree.insert("", "end", values=row, tags=(tag,))
        tree.tag_configure("even", background=COLORS["card"])
        tree.tag_configure("odd", background=COLORS["row_alt"])

        return tree

    # ─── DASHBOARD ───
    def show_dashboard(self):
        self.clear_content()
        self.page_title.config(text="🏠 Dashboard")
        
        # Pandas Analiz Məlumatlarını Alaq
        stats = get_analytics_data()
        products = load_products()
        movements = load_movements()

        if stats:
            total_products = stats["product_count"]
            total_stock_value = stats["total_stock_value"]
            potential_profit = stats["potential_profit"]
        else:
            # Fallback (Stats yoxdursa köhnə üsul) - Daha təhlükəsiz tip çevirmələri ilə
            def safe_num(val):
                try: return float(val or 0)
                except: return 0.0

            total_products = len(products)
            total_stock_value = sum(safe_num(p[7]) for p in products)
            
            # Potensial Qazanc = Əldə olan * (Satış - Alış)
            potential_profit = sum(safe_num(p[3]) * (safe_num(p[5]) - safe_num(p[4])) for p in products)

        def safe_int(val):
            try: return int(val)
            except: return 0
            
        critical_count = sum(1 for p in products if safe_int(p[3]) <= 5) # 5-dən az olanlar kritik
        today = datetime.now().strftime("%d.%m.%Y")
        today_moves = sum(1 for m in movements if m[1] and str(m[1]).startswith(today))

        # Cards Layout
        cards_frame = tk.Frame(self.content, bg=COLORS["bg"])
        cards_frame.pack(fill="x", pady=(0, 15))
        for c in range(4):
            cards_frame.columnconfigure(c, weight=1)

        card_data = [
            ("Ümumi Məhsul", total_products, COLORS["accent"], "📦"),
            ("Stok Dəyəri (Maya)", f"₼{total_stock_value:,.2f}", COLORS["warning"], "🏷️"),
            ("Potensial Qazanc", f"₼{potential_profit:,.2f}", COLORS["accent2"], "📈"),
            ("Kritik Stok", critical_count, COLORS["danger"], "⚠️"),
        ]
        for i, (title, val, color, icon) in enumerate(card_data):
            card = self.make_card(cards_frame, title, val, color, icon)
            card.grid(row=0, column=i, padx=5, sticky="nsew")

        # Main Dashboard Content (Split)
        main_container = tk.Frame(self.content, bg=COLORS["bg"])
        main_container.pack(fill="both", expand=True)
        
        # Left Side: Charts
        left_panel = tk.Frame(main_container, bg=COLORS["bg"])
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Sales Chart
        chart_card = tk.Frame(left_panel, bg=COLORS["card"], padx=10, pady=10)
        chart_card.pack(fill="both", expand=True, pady=(0, 10))
        tk.Label(chart_card, text="📈 Satış Dinamikası", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.draw_sales_chart(chart_card, movements)

        # Right Side: Top Products & Recent Actions
        right_panel = tk.Frame(main_container, bg=COLORS["bg"], width=400)
        right_panel.pack(side="right", fill="both")

        # Top Products Card
        top_card = tk.Frame(right_panel, bg=COLORS["card"], padx=15, pady=15)
        top_card.pack(fill="both", expand=True, pady=(0, 10))
        tk.Label(top_card, text="🏆 Ən Çox Satılanlar", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 10))
        
        if stats and stats["top_selling"]:
            for name, qty in stats["top_selling"]:
                row = tk.Frame(top_card, bg=COLORS["card"], pady=5)
                row.pack(fill="x")
                tk.Label(row, text=name, bg=COLORS["card"], fg=COLORS["text"], font=("Segoe UI", 9)).pack(side="left")
                tk.Label(row, text=f"{int(qty)} ədəd", bg=COLORS["card"], fg=COLORS["accent2"], font=("Segoe UI", 9, "bold")).pack(side="right")
                tk.Frame(top_card, bg=COLORS["border"], height=1).pack(fill="x")
        else:
            tk.Label(top_card, text="Məlumat yoxdur", bg=COLORS["card"], fg=COLORS["text_dim"]).pack(pady=20)

        # Recent Moves
        moves_card = tk.Frame(right_panel, bg=COLORS["card"], padx=15, pady=15)
        moves_card.pack(fill="both", expand=True)
        tk.Label(moves_card, text="🕒 Son Hərəkətlər", bg=COLORS["card"],
                 fg=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 10))
        
        recent = movements[-5:][::-1]
        for m in recent:
            color = COLORS["accent2"] if m[3] == "Giriş" else COLORS["danger"]
            row = tk.Frame(moves_card, bg=COLORS["card"], pady=4)
            row.pack(fill="x")
            tk.Label(row, text=m[2], bg=COLORS["card"], fg=COLORS["text"], font=("Segoe UI", 9)).pack(side="left")
            tk.Label(row, text=f"{m[3]} ({m[4]})", bg=COLORS["card"], fg=color, font=("Segoe UI", 8, "bold")).pack(side="right")

    def draw_sales_chart(self, parent, movements):
        stats = get_analytics_data()
        if not stats or not stats.get("monthly_report"):
            tk.Label(parent, text="Məlumat yoxdur", bg=COLORS["card"], fg=COLORS["text_dim"]).pack(expand=True)
            return

        report = stats["monthly_report"]
        sorted_months = sorted(report.keys())
        
        labels = [datetime.strptime(m, "%Y-%m").strftime("%b %y") for m in sorted_months]
        income_values = [report[m]["income"] for m in sorted_months]
        expense_values = [report[m]["expense"] for m in sorted_months]

        fig, ax = plt.subplots(figsize=(5, 3), dpi=100)
        fig.patch.set_facecolor(COLORS["card"])
        ax.set_facecolor(COLORS["card"])
        
        ax.plot(labels, income_values, marker='o', color=COLORS["accent2"], linewidth=2, label="Gəlir")
        ax.plot(labels, expense_values, marker='s', color=COLORS["danger"], linewidth=2, label="Xərc")
        ax.fill_between(labels, income_values, color=COLORS["accent2"], alpha=0.1)
        
        ax.tick_params(axis='x', colors=COLORS["text_dim"], labelsize=8)
        ax.tick_params(axis='y', colors=COLORS["text_dim"], labelsize=8)
        ax.legend(facecolor=COLORS["card"], edgecolor=COLORS["border"], labelcolor=COLORS["text"], fontsize=8)
        
        for spine in ax.spines.values():
            spine.set_edgecolor(COLORS["border"])
        
        ax.grid(True, linestyle='--', alpha=0.2, color=COLORS["border"])
        
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)
        plt.close(fig)

    # ─── MƏHSUL SİYAHISI ───
    def show_products(self):
        self.clear_content()
        self.page_title.config(text="📋 Məhsul Siyahısı")

        # Top Toolbar
        toolbar = tk.Frame(self.content, bg=COLORS["bg"], pady=10)
        toolbar.pack(fill="x")

        # Search section in toolbar
        search_frame = tk.Frame(toolbar, bg=COLORS["bg"])
        search_frame.pack(side="left")
        
        tk.Label(search_frame, text="🔍", bg=COLORS["bg"], fg=COLORS["text"],
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                                bg=COLORS["card"], fg=COLORS["text"],
                                insertbackground=COLORS["text"],
                                font=("Segoe UI", 10), bd=0, width=30)
        search_entry.pack(side="left", ipady=8, padx=5)

        # Action Buttons in toolbar
        btn_frame = tk.Frame(toolbar, bg=COLORS["bg"])
        btn_frame.pack(side="right")

        tk.Button(btn_frame, text="➕ Yeni Məhsul", bg=COLORS["accent2"], fg="white",
                  font=("Segoe UI", 10, "bold"), bd=0, padx=15, pady=8,
                  cursor="hand2", command=self.show_add_product).pack(side="left", padx=5)

        tk.Button(btn_frame, text="🛒 Satış Et", bg=COLORS["danger"], fg="white",
                  font=("Segoe UI", 10, "bold"), bd=0, padx=15, pady=8,
                  cursor="hand2", command=lambda: self.show_movement("Satış")).pack(side="left", padx=5)

        # Table Section
        products = load_products()
        cols = ["Məhsul", "Gəliş sayı", "Satılan", "Əldə olan",
                "Alış qiyməti", "Satış qiyməti", "Gəlir", "Maya dərəri", "Son Yeniləmə"]
        
        self.product_tree_frame = tk.Frame(self.content, bg=COLORS["bg"])
        self.product_tree_frame.pack(fill="both", expand=True, pady=10)

        def fmt(val):
            if isinstance(val, (float, int)):
                return f"₼{val:,.2f}"
            return val or "—"

        def show_rows(data):
            for w in self.product_tree_frame.winfo_children():
                w.destroy()
            rows = []
            for p in data:
                rows.append([
                    p[0], p[1], p[2], p[3],
                    fmt(p[4]), fmt(p[5]),
                    fmt(p[6]), fmt(p[7]),
                    p[8]
                ])
            tree = self.make_table(self.product_tree_frame, cols, rows,
                                   [180, 100, 90, 100, 130, 130, 110, 120, 150])
            tree.bind("<Double-1>", lambda e: self.delete_selected(tree, products))

        def filter_products(*_):
            q = self.search_var.get().lower()
            filtered = [p for p in products if q in str(p[0]).lower()]
            show_rows(filtered)

        self.search_var.trace("w", filter_products)
        show_rows(products)

        tk.Label(self.content, text="💡 Silmək üçün sətirə iki dəfə klikləyin",
                 bg=COLORS["bg"], fg=COLORS["text_dim"], font=("Segoe UI", 8)).pack(anchor="e")

    def delete_selected(self, tree, products):
        selected = tree.selection()
        if not selected:
            return
        vals = tree.item(selected[0])["values"]
        name = vals[0]
        if messagebox.askyesno("Sil", f"'{name}' məhsulunu silmək istədiyinizdən əminsiniz?"):
            delete_product(name)
            self.show_products()

    # ─── MƏHSUL ƏLAVƏ ET ───
    def show_add_product(self):
        self.clear_content()
        self.page_title.config(text="➕ Yeni Məhsul Əlavə et")

        form = tk.Frame(self.content, bg=COLORS["card"], padx=30, pady=25)
        form.pack(fill="x", padx=50, pady=20)

        tk.Label(form, text="Yeni Məhsul Məlumatları", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Segoe UI", 13, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))

        field_labels = [
            "Məhsul adı *",
            "Gəliş sayı *",
            "Satılan *",
            "Alış qiyməti (₼) *",
            "Satış qiyməti (₼) *",
        ]
        field_keys = ["ad", "gelis", "satilan", "alis", "satis"]
        fields = {}

        for i, (label, key) in enumerate(zip(field_labels, field_keys)):
            tk.Label(form, text=label, bg=COLORS["card"], fg=COLORS["text_dim"],
                     font=("Segoe UI", 10)).grid(row=i+1, column=0, sticky="w", pady=6, padx=(0, 20))
            var = tk.StringVar()
            entry = tk.Entry(form, textvariable=var, bg=COLORS["bg"], fg=COLORS["text"],
                             insertbackground=COLORS["text"], font=("Segoe UI", 10),
                             bd=0, width=30)
            entry.grid(row=i+1, column=1, sticky="ew", pady=6, ipady=6)
            fields[key] = var

        def save():
            ad = fields["ad"].get().strip()
            if not ad:
                messagebox.showerror("Xəta", "Məhsul adı daxil edin!")
                return
            try:
                gelis = int(fields["gelis"].get())
                satilan = int(fields["satilan"].get())
                alis = float(fields["alis"].get())
                satis = float(fields["satis"].get())
            except ValueError:
                messagebox.showerror("Xəta", "Rəqəmsal sahələri düzgün daxil edin!")
                return

            elde = gelis - satilan
            if elde < 0:
                messagebox.showerror("Xəta", "Satılan say gəliş sayından çox ola bilməz!")
                return

            now = datetime.now().strftime("%d.%m.%Y %H:%M")
            # gelir and maya will be set as formulas in save_product
            save_product([ad, gelis, satilan, elde, alis, satis, None, None, now])
            messagebox.showinfo("Uğurlu", f"✅ Məhsul əlavə edildi: {ad}")
            self.show_products()

        btn_frame = tk.Frame(form, bg=COLORS["card"])
        btn_frame.grid(row=len(field_labels)+1, column=0, columnspan=2, pady=(20, 0))
        tk.Button(btn_frame, text="✅  Saxla", bg=COLORS["accent2"], fg="white",
                  font=("Segoe UI", 11, "bold"), bd=0, padx=25, pady=10,
                  cursor="hand2", command=save).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✖  Ləğv et", bg=COLORS["danger"], fg="white",
                  font=("Segoe UI", 11), bd=0, padx=25, pady=10,
                  cursor="hand2", command=self.show_products).pack(side="left", padx=5)

    # ─── STOK GİRİŞ/ÇIXIŞ ───
    def show_movement(self, hareket_turu):
        self.clear_content()
        icon = "🛒" if hareket_turu == "Satış" else "📥"
        label_az = "Satış" if hareket_turu == "Satış" else "Giriş"
        self.page_title.config(text=f"{icon} {label_az} Əməliyyatı")
        color = COLORS["accent2"] if hareket_turu == "Giriş" else COLORS["danger"]

        products = load_products()
        product_map = {p[0]: p for p in products}
        product_options = [p[0] for p in products]

        form = tk.Frame(self.content, bg=COLORS["card"], padx=30, pady=25)
        form.pack(fill="x", padx=50, pady=20)

        tk.Label(form, text=f"Stok {label_az} Əməliyyatı", bg=COLORS["card"],
                 fg="white", font=("Segoe UI", 13, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))

        tk.Label(form, text="Məhsul Axtar", bg=COLORS["card"], fg=COLORS["text_dim"],
                 font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=8, padx=(0, 20))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(form, textvariable=search_var, bg=COLORS["bg"], fg=COLORS["text"],
                                insertbackground=COLORS["text"], font=("Segoe UI", 10), bd=0, width=35)
        search_entry.grid(row=1, column=1, sticky="ew", pady=8, ipady=6)
        
        tk.Label(form, text="Məhsul Seçin *", bg=COLORS["card"], fg=COLORS["text_dim"],
                 font=("Segoe UI", 10)).grid(row=2, column=0, sticky="w", pady=8, padx=(0, 20))
        
        product_var = tk.StringVar()
        product_cb = ttk.Combobox(form, textvariable=product_var, values=product_options,
                                   state="readonly", width=35, font=("Segoe UI", 10))
        product_cb.grid(row=2, column=1, sticky="ew", pady=8)

        info_label = tk.Label(form, text="", bg=COLORS["card"], fg=COLORS["accent"],
                                font=("Segoe UI", 9, "italic"))
        info_label.grid(row=3, column=1, sticky="w")

        def update_search(*_):
            q = search_var.get().lower()
            filtered = [p for p in product_options if q in p.lower()]
            product_cb["values"] = filtered
            if filtered:
                product_cb.set(filtered[0])
            else:
                product_cb.set("")

        search_var.trace("w", update_search)

        def on_product_select(*_):
            sel = product_var.get()
            if sel and sel in product_map:
                p = product_map[sel]
                info_label.config(text=f"Əldə olan: {p[3]}  |  Satılan: {p[2]}  |  Gəliş: {p[1]}")

        product_var.trace("w", on_product_select)

        tk.Label(form, text="Miqdar *", bg=COLORS["card"], fg=COLORS["text_dim"],
                 font=("Segoe UI", 10)).grid(row=4, column=0, sticky="w", pady=8, padx=(0, 20))
        miktar_var = tk.StringVar()
        tk.Entry(form, textvariable=miktar_var, bg=COLORS["bg"], fg=COLORS["text"],
                 insertbackground=COLORS["text"], font=("Segoe UI", 10), bd=0, width=30
                 ).grid(row=4, column=1, sticky="ew", pady=8, ipady=6)

        tk.Label(form, text="Açıqlama", bg=COLORS["card"], fg=COLORS["text_dim"],
                 font=("Segoe UI", 10)).grid(row=5, column=0, sticky="w", pady=8, padx=(0, 20))
        aciklama_var = tk.StringVar()
        tk.Entry(form, textvariable=aciklama_var, bg=COLORS["bg"], fg=COLORS["text"],
                 insertbackground=COLORS["text"], font=("Segoe UI", 10), bd=0, width=30
                 ).grid(row=5, column=1, sticky="ew", pady=8, ipady=6)

        def process():
            sel = product_var.get()
            if not sel:
                messagebox.showerror("Xəta", "Məhsul seçin!")
                return
            try:
                miktar = int(miktar_var.get())
                if miktar <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Xəta", "Düzgün miqdar daxil edin!")
                return

            delta = miktar if hareket_turu == "Giriş" else -miktar
            ok, msg = update_product_stock(sel, delta, hareket_turu,
                                           aciklama_var.get() or f"{label_az} əməliyyatı",
                                           self.current_user)
            if ok:
                messagebox.showinfo("Uğurlu", f"✅ {msg}")
                self.navigate("dashboard")
            else:
                messagebox.showerror("Xəta", msg)

        btn_frame = tk.Frame(form, bg=COLORS["card"])
        btn_frame.grid(row=6, column=0, columnspan=2, pady=(20, 0))
        tk.Button(btn_frame, text=f"{icon}  {label_az}", bg=color, fg="white",
                  font=("Segoe UI", 11, "bold"), bd=0, padx=25, pady=10,
                  cursor="hand2", command=process).pack(side="left", padx=5)
        tk.Button(btn_frame, text="✖  Ləğv et", bg=COLORS["border"], fg="white",
                  font=("Segoe UI", 11), bd=0, padx=25, pady=10,
                  cursor="hand2", command=self.show_dashboard).pack(side="left", padx=5)

    # ─── HƏRƏKƏTLƏr ───
    def show_movements(self):
        self.clear_content()
        self.page_title.config(text="📊 Stok Hərəkətləri")
        movements = load_movements()
        cols = ["Hərəkət ID", "Tarix", "Məhsul", "Növ", "Miqdar", "Açıqlama", "Qeyd edən"]
        rows = [[m[0], m[1], m[2], m[3], m[4], m[5], m[6]] for m in movements[::-1]]
        self.make_table(self.content, cols, rows, [100, 140, 200, 90, 80, 260, 110])

    # ─── KRİTİK STOK ───
    def show_critical(self):
        self.clear_content()
        self.page_title.config(text="⚠️ Kritik Stok Vəziyyəti")
        products = load_products()
        # Critical = əldə olan <= 0 (ensure numeric)
        def safe_int(val):
            try:
                return int(float(val))
            except Exception:
                return 0
        critical = [p for p in products if safe_int(p[3]) <= 0]

        if not critical:
            tk.Label(self.content, text="✅ Kritik stok vəziyyəti yoxdur!",
                     bg=COLORS["bg"], fg=COLORS["accent2"],
                     font=("Segoe UI", 16)).pack(expand=True)
            return

        tk.Label(self.content, text=f"⚠️  {len(critical)} məhsul kritik stok səviyyəsindədir!",
                 bg=COLORS["bg"], fg=COLORS["warning"],
                 font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 10))

        cols = ["Məhsul", "Gəliş sayı", "Satılan", "Əldə olan", "Alış qiyməti", "Satış qiyməti"]
        rows = [[p[0], p[1], p[2], p[3], p[4], p[5]] for p in critical]
        self.make_table(self.content, cols, rows, [200, 110, 100, 110, 140, 140])


    # ─── MALİYYƏ HESABATI ───
    def show_financials(self):
        self.clear_content()
        self.page_title.config(text="💰 Aylıq Maliyyə Hesabatı")

        stats = get_analytics_data()
        if not stats or not stats.get("monthly_report"):
            tk.Label(self.content, text="Hesabat üçün yetərli məlumat yoxdur.",
                     bg=COLORS["bg"], fg=COLORS["text_dim"], font=("Segoe UI", 12)).pack(expand=True)
            return

        report = stats["monthly_report"]
        cols = ["Ay", "Ümumi Gəlir (₼)", "Ümumi Xərc (₼)", "Xalis Mənfəət (₼)"]
        rows = []
        
        for m_key in sorted(report.keys(), reverse=True):
            data = report[m_key]
            inc = data["income"]
            exp = data["expense"]
            prof = inc - exp
            rows.append([
                datetime.strptime(m_key, "%Y-%m").strftime("%B %Y"),
                f"₼{inc:,.2f}",
                f"₼{exp:,.2f}",
                f"₼{prof:,.2f}"
            ])

        self.make_table(self.content, cols, rows, [200, 180, 180, 180])

    # ─── ANALİZ ───
    def show_analysis(self):
        self.clear_content()
        self.page_title.config(text="📈 Analiz Mərkəzi")
        self.analysis_filter = getattr(self, "analysis_filter", 1)  # default 1 month

        # ── Filter Toolbar ──
        filter_bar = tk.Frame(self.content, bg=COLORS["bg"], pady=8)
        filter_bar.pack(fill="x")

        tk.Label(filter_bar, text="Dövr:", bg=COLORS["bg"], fg=COLORS["text"],
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=(0, 10))

        filter_options = [("Aylıq", 1), ("3 Aylıq", 3), ("6 Aylıq", 6), ("İllik", 12)]
        self._filter_btns = {}
        for label, val in filter_options:
            is_active = (val == self.analysis_filter)
            bg = COLORS["accent"] if is_active else COLORS["card"]
            fg = "white" if is_active else COLORS["text"]
            btn = tk.Button(
                filter_bar, text=label,
                bg=bg, fg=fg,
                font=("Segoe UI", 9, "bold" if is_active else "normal"),
                bd=0, padx=14, pady=6, cursor="hand2",
                command=lambda v=val: self._apply_analysis_filter(v)
            )
            btn.pack(side="left", padx=4)
            self._filter_btns[val] = btn

        # ── Fetch Data ──
        data = get_filtered_analysis_data(self.analysis_filter)

        if not data:
            tk.Label(self.content, text="Məlumat yoxdur. Əvvəlcə satış əməliyyatları aparın.",
                     bg=COLORS["bg"], fg=COLORS["text_dim"],
                     font=("Segoe UI", 13)).pack(expand=True)
            return

        income  = data["total_income"]
        expense = data["total_expense"]
        profit  = income - expense

        # ── Summary Cards ──
        cards_frame = tk.Frame(self.content, bg=COLORS["bg"])
        cards_frame.pack(fill="x", pady=(0, 12))
        for c in range(3):
            cards_frame.columnconfigure(c, weight=1)

        card_data = [
            ("Ümumi Gəlir",   f"₼{income:,.2f}",  COLORS["accent2"], "💰"),
            ("Ümumi Xərc",    f"₼{expense:,.2f}", COLORS["danger"],  "🏷️"),
            ("Xalis Mənfəət", f"₼{profit:,.2f}",  COLORS["accent"],  "📈"),
        ]
        for i, (title, val, color, icon) in enumerate(card_data):
            card = self.make_card(cards_frame, title, val, color, icon)
            card.grid(row=0, column=i, padx=6, sticky="nsew")

        # ── Charts Row ──
        charts_row = tk.Frame(self.content, bg=COLORS["bg"])
        charts_row.pack(fill="both", expand=True)
        charts_row.columnconfigure(0, weight=3)
        charts_row.columnconfigure(1, weight=2)

        # Left chart: Monthly income vs expense
        left_card = tk.Frame(charts_row, bg=COLORS["card"], padx=10, pady=10)
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        tk.Label(left_card, text="📊 Aylıq Gəlir / Xərc Dinamikası",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

        monthly = data["monthly"]
        if monthly:
            sorted_months = sorted(monthly.keys())
            labels    = [datetime.strptime(m, "%Y-%m").strftime("%b %Y") for m in sorted_months]
            incomes   = [monthly[m]["income"]  for m in sorted_months]
            expenses  = [monthly[m]["expense"] for m in sorted_months]
            profits   = [i - e for i, e in zip(incomes, expenses)]

            fig1, ax1 = plt.subplots(figsize=(6, 3), dpi=100)
            fig1.patch.set_facecolor(COLORS["card"])
            ax1.set_facecolor(COLORS["card"])

            x = range(len(labels))
            width = 0.28
            ax1.bar([i - width for i in x], incomes,  width=width, color=COLORS["accent2"], label="Gəlir",   alpha=0.85)
            ax1.bar(x,                      expenses, width=width, color=COLORS["danger"],  label="Xərc",    alpha=0.85)
            ax1.bar([i + width for i in x], profits,  width=width, color=COLORS["accent"],  label="Mənfəət", alpha=0.85)

            ax1.set_xticks(list(x))
            ax1.set_xticklabels(labels, rotation=30, ha="right", fontsize=7, color=COLORS["text_dim"])
            ax1.tick_params(axis='y', colors=COLORS["text_dim"], labelsize=7)
            ax1.legend(facecolor=COLORS["card"], edgecolor=COLORS["border"],
                       labelcolor=COLORS["text"], fontsize=7)
            for spine in ax1.spines.values():
                spine.set_edgecolor(COLORS["border"])
            ax1.grid(axis='y', linestyle='--', alpha=0.2, color=COLORS["border"])

            canvas1 = FigureCanvasTkAgg(fig1, master=left_card)
            canvas1.draw()
            canvas1.get_tk_widget().pack(fill="both", expand=True)
            plt.close(fig1)
        else:
            tk.Label(left_card, text="Bu dövrdə məlumat yoxdur.",
                     bg=COLORS["card"], fg=COLORS["text_dim"],
                     font=("Segoe UI", 10)).pack(expand=True)

        # Right chart: Top product profitability (horizontal bar)
        right_card = tk.Frame(charts_row, bg=COLORS["card"], padx=10, pady=10)
        right_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        tk.Label(right_card, text="🏆 Ən Çox Qazandıran Məhsullar",
                 bg=COLORS["card"], fg=COLORS["text"],
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

        product_profit = data["product_profit"]
        if product_profit:
            prods   = [p[0][:18] for p in product_profit]
            profits2 = [p[1] for p in product_profit]
            colors  = [COLORS["accent2"] if v >= 0 else COLORS["danger"] for v in profits2]

            fig2, ax2 = plt.subplots(figsize=(4, 3), dpi=100)
            fig2.patch.set_facecolor(COLORS["card"])
            ax2.set_facecolor(COLORS["card"])

            bars = ax2.barh(prods[::-1], profits2[::-1], color=colors[::-1], alpha=0.85)
            ax2.tick_params(axis='x', colors=COLORS["text_dim"], labelsize=7)
            ax2.tick_params(axis='y', colors=COLORS["text"],    labelsize=7)
            for spine in ax2.spines.values():
                spine.set_edgecolor(COLORS["border"])
            ax2.grid(axis='x', linestyle='--', alpha=0.2, color=COLORS["border"])
            ax2.axvline(0, color=COLORS["border"], linewidth=0.8)

            # Value labels on bars
            for bar, val in zip(bars, profits2[::-1]):
                ax2.text(
                    bar.get_width() + (abs(max(profits2, default=1)) * 0.02),
                    bar.get_y() + bar.get_height() / 2,
                    f"₼{val:,.0f}",
                    va='center', fontsize=6,
                    color=COLORS["text_dim"]
                )

            fig2.tight_layout()
            canvas2 = FigureCanvasTkAgg(fig2, master=right_card)
            canvas2.draw()
            canvas2.get_tk_widget().pack(fill="both", expand=True)
            plt.close(fig2)
        else:
            tk.Label(right_card, text="Bu dövrdə satış məlumatı yoxdur.",
                     bg=COLORS["card"], fg=COLORS["text_dim"],
                     font=("Segoe UI", 10)).pack(expand=True)

    def _apply_analysis_filter(self, months):
        self.analysis_filter = months
        self.show_analysis()


if __name__ == "__main__":
    root = tk.Tk()
    app = DepoApp(root)
    root.mainloop()